from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.export.excel.base_renderer import BaseExcelRenderer
from app.services.export.core.render_context import RenderContext

class TableExcelRenderer(BaseExcelRenderer):
    
    def build_document(self, wb: Workbook, ws, dataset: Dict[str, Any], context: RenderContext):
        metadata = dataset.get("metadata", {})
        summary = dataset.get("summary", {})
        
        # Meta info
        ws.append([metadata.get("company", "Company")])
        ws.append([f"Report: {metadata.get('report_type', 'Export')}"])
        ws.append([f"Period: {metadata.get('period', 'All Time')}"])
        ws.append([])
        
        # Format headers
        font_bold = Font(bold=True)
        for row in range(1, 4):
            ws.cell(row=row, column=1).font = font_bold
            
        headers = dataset.get("headers", [])
        rows = dataset.get("rows", [])
        
        if not headers and not rows and "activities" in dataset:
            headers = ["Date", "Activity", "Employee", "Division", "Category", "Status", "Duration", "Notes"]
            for act in dataset["activities"]:
                rows.append([
                    act.get("date"),
                    act.get("activity"),
                    act.get("employee"),
                    act.get("division"),
                    act.get("category"),
                    act.get("status"),
                    act.get("duration_human"),
                    act.get("notes")
                ])
                
        if headers:
            ws.append(headers)
            header_row_idx = ws.max_row
            
            fill_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row_idx, column=col)
                cell.font = font_bold
                cell.fill = fill_header
                
            for row in rows:
                ws.append(row)
                
            # Auto adjust columns loosely
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                if adjusted_width > 50:
                    adjusted_width = 50
                ws.column_dimensions[column].width = adjusted_width
